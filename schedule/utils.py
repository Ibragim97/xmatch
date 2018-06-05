import os
import re
import string
import random
import requests
import datetime

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill

from HTMLParser import HTMLParser
from tabula import read_pdf

from schedule_match import settings
from schedule.models import Session, School, Subject, SubjectInstance

INF = int(1e+9)

def parse_pdf(school="SHSS", fileName="shss.pdf"):

    def get_text(col):
        return col['text'].replace('\r', ' ')

    def get_time(timeString):
        date = datetime.datetime.strptime(timeString, "%H:%M %p")
        if date.hour < 9:
            return datetime.datetime(1900, 1, 1, date.hour + 12, date.minute)
        return date


    def create_subject(abbr, title, credits):
        if abbr:
            try:
                subject = Subject.objects.get(abbr=abbr)
            except Subject.DoesNotExist:
                subject = Subject(school=schoolObj, abbr=abbr, title=title, credits=credits)
                subject.save()
            return subject
        return None

    def create_subInst(subject, section, days, time, faculty, room):
        start, end = time.split('-')

        inst = SubjectInstance(subject=subject, section=section, days=days, 
                               startTime=get_time(start), endTime=get_time(end),
                               faculty=faculty, room=room)
        inst.save()
        return

    print "Started parsing %s" % school
        
    schoolObj = School.objects.get(name=school)
    headers = ["Course Abbr", 'S/T', "Course Title", "Cr(US)", 
                "Days", "Time", "Faculty", "Room"]
    ind = []

    data = read_pdf(fileName, spreadsheet=True, pages=1, output_format="json")
    row = data[0]['data'][0]

    for i, col in enumerate(row):
        if get_text(col) in headers:
            ind.append(i)

    for page in range(1, 100):
        try:
            data = read_pdf(fileName, spreadsheet=True, pages=page, output_format="json")
        except:
            break
        else:
            try:
                rows = data[0]['data']
            except:
                break
            if page == 1:
                rows = rows[1:]

            for row in rows:
                abbr = get_text(row[ind[0]])
                section = get_text(row[ind[1]])
                title = get_text(row[ind[2]])
                credits = get_text(row[ind[3]])
                days = get_text(row[ind[4]])
                time = get_text(row[ind[5]])
                faculty = get_text(row[ind[6]])
                room = get_text(row[ind[7]])

                subject = create_subject(abbr, title, credits)
                if subject:
                    create_subInst(subject, section, days, time, faculty, room)
    return

def refresh_data():
    SEMESTR = "Spring 2018"
    SST = "The School of Science and Technology"
    SHSS = "The School of Humanities and Social Sciences"
    REGISTRAR_URL = "https://registrar.nu.edu.kz/course-schedules"

    def create_pdf(data, shortName, lastModified, fileName):
        school = School(name=shortName, last_modified=lastModified)
        school.save()
        print "Creating pdf for %s" % shortName
        with open(fileName, 'wb') as file:
            file.write(data)

        parse_pdf(shortName, fileName)

    def download_file(page, shortName, school, fileName):

        if shortName == "SST":
            school_url = "http://registrar.nu.edu.kz/registrar_downloads/json?method=printDocument&name=school_schedule_by_term&termid=301&schoolid=2"
        else:
            school_url = "http://registrar.nu.edu.kz/registrar_downloads/json?method=printDocument&name=school_schedule_by_term&termid=301&schoolid=3"            

        if not school_url:
            rows = page.split(SEMESTR, 1)[1].splitlines()
            mark = True
            h = HTMLParser()
            print "Started downloading %s" % shortName
            for row in rows:
                if mark and school in row:
                    school_url = h.unescape(row.split('<a')[1])
                    school_url = school_url.split('"')[1]
                    mark = False

        request = requests.get(school_url)
        newDate = request.headers['Last-Modified']

        try:
            obj = School.objects.get(name=shortName)
        except:
            create_pdf(request.content, shortName, newDate, fileName)
        else:
            prevDate = obj.last_modified
            if prevDate != newDate:
                obj.delete()
                create_pdf(request.content, shortName, newDate, fileName)

        return 

    try:
        page = requests.get(REGISTRAR_URL).content
    except:
        print "Error to get registrar page"
    else:
        download_file(page, "SST", SST, 'sst.pdf')
        print "SST Done"
        download_file(page, "SHSS", SHSS, 'shss.pdf')


required = ["MATH 162", "MATH 273", "PHYS 161", "HST 100"]
# required = ["PHYS 161"]

def generate(hashcode="1", courses=required, optional=[]):

    def splitByType(sections):
        result = []
        different = []

        for inst in sections:
            sectionType = re.findall("[a-zA-z]+", inst.section)

            if sectionType not in different:
                different.append(sectionType)
                result.append([inst])
            else:
                i = different.index(sectionType)
                result[i].append(inst)

        return result

    def gen_combination(combinations, pos):
        if pos == sub_len:
            combinations.append(arr[:])
            return

        for val in range(len(subjects[pos])):
            arr[pos] = val
            gen_combination(combinations, pos+1) 

    def rate(combinations):
        daysWeek = ['M', 'T', 'W', 'R', 'F', 'S', '']

        for combID, comb in enumerate(combinations):
            week = [[] for i in range(7)]
            for i, pos in enumerate(comb):
                instance = subjects[i][pos]
                days = instance.days.split(' ')
                for day in days:
                    week[daysWeek.index(day)].append((instance.startTime, instance.endTime))

            rating = 0
            fine = 0

            for i in range(7):
                week[i].sort()

                for j, sub in enumerate(week[i]):
                    if j == len(week[i]) - 1:
                        break
                    dist = (week[i][j+1][0] - week[i][j][1]).total_seconds() / 60

                    if dist < 0:
                        rating = INF
                        break
                    
                    rating += 100 * fine * dist
                    
                    if dist > 40:
                        fine += 1

                if rating == INF:
                    break

            ratings[combID] = (rating, combID)

        return

    def exportCombinations():
        HEAD_HIGHT = 45
        daysWeek = ['M', 'T', 'W', 'R', 'F', 'S']
        daysWeekF = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        colors = []

        def gen_randFileName():
            return ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(6))

        def formatedTime(inst):
            return "%s-%s" % (inst.startTime.strftime("%H:%M"), 
                              inst.endTime.strftime("%H:%M"))

        def formatedText1(inst):
            return "%s (%s) \n%s \n%s \n%s" % (inst.subject.title, inst.section,   
                                               formatedTime(inst), inst.faculty, inst.room)
        def formatedText2(inst):
            return "%s (%s) \n%s" % (inst.subject.title, inst.section, formatedTime(inst))            

        def initWSstyles():
            colors.append(PatternFill("solid", fgColor="2c8df4")) #blue
            colors.append(PatternFill("solid", fgColor="17ed45")) #green
            colors.append(PatternFill("solid", fgColor="f43f3f")) #red
            colors.append(PatternFill("solid", fgColor="fcff72")) #yellow
            colors.append(PatternFill("solid", fgColor="d8efff")) #white
            colors.append(PatternFill("solid", fgColor="f4892c")) #orange
            
            ws.column_dimensions['A'].width = 5

            for i in range(5):
                ws.column_dimensions[chr(ord('B')+i)].width = 21
            ws.column_dimensions['G'].width = 15

            normal = NamedStyle(name="normal")
            normal.font = Font(size=9)
            normal.alignment.vertical = "top"
            normal.alignment.wrap_text = True
            wb.add_named_style(normal)

        def writeSubject(inst, row):
            rowCh = str(row+2)
            ws['H1'] = "Abbr"
            ws['H'+rowCh] = instance.subject.abbr
            ws.column_dimensions['H'].width = 10

            ws['I1'] = "Title"
            ws['I'+rowCh] = instance.subject.title
            ws.column_dimensions['I'].width = 25

            ws['J1'] = "S"
            ws['J'+rowCh] = instance.section
            ws.column_dimensions['J'].width = 6

            ch = 'K'
            ws[ch+'1'] = "Cr(US)"
            ws[ch+rowCh] = instance.subject.credits
            ws.column_dimensions[ch].width = 6

            ch = 'L'
            ws[ch+'1'] = "Days"
            ws[ch+rowCh] = instance.days
            ws.column_dimensions[ch].width = 8

            ch = 'M'
            ws[ch+'1'] = "Time"
            ws[ch+rowCh] = formatedTime(instance)
            ws.column_dimensions[ch].width = 15

            ch = 'N'
            ws[ch+'1'] = "Faculty"
            ws[ch+rowCh] = instance.faculty
            ws.column_dimensions[ch].width = 20

            ch = 'O'
            ws[ch+'1'] = "Room"
            ws[ch+rowCh] = instance.room
            ws.column_dimensions[ch].width = 20
            
            for i in range(8):
                ws[chr(ord('H')+i)+rowCh].fill = colors[courses.index(instance.subject.abbr)]

        ids = [x[1] for x in ratings[:settings.NUM_OF_TABLES]]
        files = []
        tables = [[] for i in range(settings.NUM_OF_TABLES)]

        for part, idx in enumerate(ids):
            if ratings[part][0] == INF:
                break
            comb = combinations[idx]
            # print comb

            wb = Workbook()
            ws = wb.active
            initWSstyles()

            week = [[] for i in range(6)]
        
            for i, pos in enumerate(comb):
                instance = subjects[i][pos]
                tables[part].append(instance)
                days = instance.days.split(' ')
                for day in days:
                    week[daysWeek.index(day)].append(instance)
                # print "%s %s" % (instance.subject.title, instance.section)
                #print table of subjects
                writeSubject(instance, i)

            # Init time column
            time = datetime.datetime(1900, 1, 1, 9, 0)
            timeRow = [None, None]
            for i in range(2, 100):
                ws['A'+str(i)] = time.time().strftime("%H:%M")
                ws['A'+str(i)].style = 'normal'
                timeRow.append(time.time())
                if time.hour > 19:
                    break
                time += datetime.timedelta(minutes=30)

            # Init days of the week row
            for i, day in enumerate(daysWeekF):
                ch =chr(ord('B') + i)
                ws[ch+'1'] = day

            # fill table with SubjectInstances
            for dayCol, day in enumerate(week):
                # print daysWeekF[dayCol]
                day.sort(key=lambda x: x.startTime, reverse=False)
                for inst in day:
                    # print "%s: %s %s - %s" % (inst, inst.days, inst.startTime.time(), inst.endTime.time())
                    cur = inst.startTime
                    rowInd = timeRow.index(cur.time())
                    first = True

                    while cur.time() < inst.endTime.time():

                        if first:
                            text = formatedText1(inst)
                            first = False
                            ws.row_dimensions[rowInd].height = HEAD_HIGHT
                        else:
                            text = formatedText2(inst)
                            if ws.row_dimensions[rowInd].height != HEAD_HIGHT:       
                                ws.row_dimensions[rowInd].height = 30       

                        
                        ch = chr(ord('B') + dayCol)
                        ws[ch+str(rowInd)] = text
                        ws[ch+str(rowInd)].style = 'normal'
                        ws[ch+str(rowInd)].fill = colors[courses.index(inst.subject.abbr)]
                        
                        rowInd += 1
                        cur += datetime.timedelta(minutes=30)
                # print ""
            fileName = "table_%s_%s.xlsx" % (hashcode, part)
            fileName = os.path.join(settings.TABLES_DIR, fileName)
            wb.save(fileName)
            files.append(fileName)

        session = Session.objects.get(hashcode=hashcode)
        session.numOfTables = len(files)
        # print len(files)
        session.save()
        return tables
            
    subjects = []
    
    for course in courses:
        subject = Subject.objects.get(abbr=course)
        sections = subject.subjectinstance_set.all()
        subjects += splitByType(sections)

    # lens = []
    # for sub in subjects:
    #     lens.append(len(sub))
    # print lens

    
    sub_len = len(subjects)
    arr = range(sub_len)    
    combinations = []
    gen_combination(combinations, 0)
    # print courses
    # print len(combinations)
    # print "GENERATED"

    # rate combinations
    # ratings has tuple(rating, id of Comb)
    ratings = range(len(combinations))
    rate(combinations)
    print "RATED"
    ratings.sort()

    return exportCombinations()


def deleteTables():
    now = datetime.datetime.now() - datetime.timedelta(minutes=30)
    sessions = Session.objects.filter(created__lte=now)
    
    for session in sessions:
        for part in range(session.numOfTables):
            fileName = "table_%s_%s.xlsx" % (session.hashcode, part)
            file_path = os.path.join(settings.TABLES_DIR, fileName)
            try:
                os.remove(file_path)
            except:
                break
        session.delete()







# def excelToPDF(path):
#     pass
#     # import excel2img
#     # # Save as BMP the range B2:C15 in test.xlsx on page named "Sheet2"
#     # excel2img.export_img("test.xlsx", "test.bmp", "", "Sheet2!B2:C15")    